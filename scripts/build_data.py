#!/usr/bin/env python3
"""
build_data.py — Convert CoStar multifamily exports into the dashboard's data blob.

Usage:
    python3 scripts/build_data.py --metro METRO.xlsx --sub SUB.csv --out data/current.js

Produces a JS file defining `var D = {...}` in the dashboard schema:
    p   list of 62 period labels
    lp  last actual period (used for the ml snapshot + forecast boundary)
    ml  metro list: snapshot of each metro at lp  [{n,c,vr,ar,rg,inv,abs,uc,sp,cr,pop,mhi}]
    mt  metro timeseries: {metro: {period: {26 base metrics + 8 derived growth metrics}}}
    st  submarket timeseries: {sub: {period: {vr,ar,rg,inv,abs,nd,uc}}}
    cm  cbsa -> metro name
    cs  cbsa -> [submarket names]
    nt  national aggregate: {period: {vr,ar,er,inv,abs,nd,uc,cs,mhi,pop,hh,emp,av}}

Also emits a small JSON sidecar (data/current.meta.json) describing the vintage.
"""
import argparse, csv, json, sys
from collections import defaultdict, OrderedDict

# ---- column -> schema-key mapping for metro timeseries (mt) ----
# (schema_key, csv_column, rounding)  rounding: 'i'=int, or number of decimals
MT_MAP = [
    ("vr",   "Vacancy Rate",                       4),
    ("ar",   "Market Asking Rent/Unit",            1),
    ("er",   "Market Effective Rent/Unit",         1),
    ("rg",   "Market Asking Rent Growth 12 Mo",    4),
    ("erg",  "Market Effective Rent Growth 12 Mo", 4),
    ("rgq",  "Market Asking Rent Growth",          4),
    ("inv",  "Inventory Units",                    'i'),
    ("abs",  "Absorption Units",                   'i'),
    ("nd",   "Net Delivered Units",                'i'),
    ("uc",   "Under Construction Units",           'i'),
    ("sp",   "Market Sale Price Per Unit",         1),
    ("cr",   "Market Cap Rate",                    4),
    ("pop",  "Population",                         'i'),
    ("hh",   "Households",                         'i'),
    ("emp",  "Total Employment",                   'i'),
    ("cs",   "Construction Starts Units",          'i'),
    ("ap",   "Absorption %",                       4),
    ("av",   "Asset Value",                        'i'),
    ("mhi",  "Median Household Income",            'i'),
    ("dem",  "Demand Units",                       'i'),
    ("spg",  "Market Sale Price Growth",           4),
    ("appr", "Appreciation Return",                4),
    ("incr", "Income Return",                      4),
    ("totr", "Total Return",                       4),
    ("oemp", "Office Employment",                  'i'),
    ("iemp", "Industrial Employment",              'i'),
]

# submarket timeseries (st) — only these 7 metrics
ST_MAP = [
    ("vr",  "Vacancy Rate",                    4),
    ("ar",  "Market Asking Rent/Unit",         1),
    ("rg",  "Market Asking Rent Growth 12 Mo", 4),
    ("inv", "Inventory Units",                 'i'),
    ("abs", "Absorption Units",                'i'),
    ("nd",  "Net Delivered Units",             'i'),
    ("uc",  "Under Construction Units",        'i'),
]

# YoY-growth derived fields: schema_key -> source base key (relative % change)
GROWTH = {"popg": "pop", "empg": "emp", "hhg": "hh",
          "invg": "inv", "mhig": "mhi", "avg": "av"}


def num(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() == "none":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def rnd(v, mode):
    if v is None:
        return None
    return int(round(v)) if mode == 'i' else round(v, mode)


def sort_periods(periods):
    def key(p):
        yr = int(p.split()[0])
        q = p.split()[1]  # Q1..Q4
        qn = int(q[1])
        # QTD/EST land between the prior quarter and the plain quarter label
        suf = 0
        if "QTD" in p:
            suf = 1
        elif "EST" in p:
            suf = 2
        return (yr, qn, suf)
    return sorted(periods, key=key)


def read_metro(path):
    """Read the metro xlsx -> mt dict + ordered period list + cbsa map."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ix = {c: i for i, c in enumerate(hdr)}
    mt = defaultdict(dict)
    cm = {}
    periods = set()
    for row in it:
        name = str(row[ix["Geography Name"]]).strip()
        per = row[ix["Period"]]
        periods.add(per)
        cbsa = str(row[ix["CBSA Code"]]).strip()
        cm[cbsa] = name
        rec = {}
        for key, col, mode in MT_MAP:
            rec[key] = rnd(num(row[ix[col]]), mode)
        # drop keys that are None to keep the blob tight
        mt[name][per] = {k: v for k, v in rec.items() if v is not None}
    return dict(mt), sort_periods(periods), cm


def read_sub(path):
    """Read the submarket csv -> st dict + cbsa->[subs] map."""
    st = defaultdict(dict)
    cs = defaultdict(list)
    seen_sub = set()
    with open(path, encoding="utf-8-sig") as fh:
        r = csv.DictReader(fh)
        for row in r:
            name = row["Geography Name"].strip()
            per = row["Period"]
            cbsa = str(row["CBSA Code"]).strip()
            if name not in seen_sub:
                seen_sub.add(name)
                cs[cbsa].append(name)
            rec = {}
            for key, col, mode in ST_MAP:
                rec[key] = rnd(num(row.get(col)), mode)
            rec = {k: v for k, v in rec.items() if v is not None}
            # keep a period only if it has inventory (matches old behavior of
            # variable-length submarket series)
            if rec.get("inv") is not None:
                st[name][per] = rec
    return dict(st), dict(cs)


def add_growth(mt, periods):
    """Add YoY growth + vacancy-change + uc% derived fields where t-4 exists."""
    pidx = {p: i for i, p in enumerate(periods)}
    for metro, series in mt.items():
        for p, rec in series.items():
            i = pidx[p]
            if i < 4:
                continue
            prevp = periods[i - 4]
            prev = series.get(prevp)
            if not prev:
                continue
            for gk, base in GROWTH.items():
                a, b = prev.get(base), rec.get(base)
                if a not in (None, 0) and b is not None:
                    rec[gk] = round((b - a) / a, 4)
            if prev.get("vr") is not None and rec.get("vr") is not None:
                rec["vrc"] = round(rec["vr"] - prev["vr"], 4)
            if rec.get("inv") and rec.get("uc") is not None:
                rec["ucpct"] = round(rec["uc"] / rec["inv"], 4)


def build_national(mt, periods):
    """Aggregate metros -> national: sum volumes, inventory-weight rates."""
    nt = {}
    for p in periods:
        sinv = sabs = snd = suc = scs = spop = shh = semp = sav = 0.0
        vac = arw = erw = mhiw = 0.0
        mhipop = 0.0
        n = 0
        for metro, series in mt.items():
            d = series.get(p)
            if not d:
                continue
            n += 1
            inv = d.get("inv") or 0
            sinv += inv
            sabs += d.get("abs") or 0
            snd += d.get("nd") or 0
            suc += d.get("uc") or 0
            scs += d.get("cs") or 0
            spop += d.get("pop") or 0
            shh += d.get("hh") or 0
            semp += d.get("emp") or 0
            sav += d.get("av") or 0
            if d.get("vr") is not None:
                vac += d["vr"] * inv
            if d.get("ar") is not None:
                arw += d["ar"] * inv
            if d.get("er") is not None:
                erw += d["er"] * inv
            if d.get("mhi") is not None and d.get("pop"):
                mhiw += d["mhi"] * d["pop"]
                mhipop += d["pop"]
        if n == 0:
            continue
        rec = {
            "vr": round(vac / sinv, 4) if sinv else None,
            "ar": round(arw / sinv, 1) if sinv else None,
            "er": round(erw / sinv, 1) if sinv else None,
            "inv": int(sinv), "abs": int(sabs), "nd": int(snd),
            "uc": int(suc), "cs": int(scs),
            "mhi": int(round(mhiw / mhipop)) if mhipop else None,
            "pop": int(spop), "hh": int(shh), "emp": int(semp),
            "av": int(sav),
        }
        nt[p] = {k: v for k, v in rec.items() if v is not None}
    return nt


def build_ml(mt, cm, lp):
    """Metro list snapshot at lp."""
    name_to_cbsa = {v: k for k, v in cm.items()}
    ml = []
    for metro, series in sorted(mt.items(), key=lambda kv: -(kv[1].get(lp, {}).get("inv") or 0)):
        d = series.get(lp, {})
        entry = OrderedDict()
        entry["n"] = metro
        cbsa = name_to_cbsa.get(metro)
        entry["c"] = int(cbsa) if cbsa and cbsa.isdigit() else cbsa
        for k in ("vr", "ar", "rg", "inv", "abs", "uc", "sp", "cr", "pop", "mhi"):
            if d.get(k) is not None:
                entry[k] = d[k]
        ml.append(entry)
    return ml


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--metro", required=True)
    ap.add_argument("--sub", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--varname", default="D")
    args = ap.parse_args()

    print("Reading metro export...", file=sys.stderr)
    mt, periods, cm = read_metro(args.metro)
    print(f"  {len(mt)} metros, {len(periods)} periods", file=sys.stderr)

    print("Reading submarket export...", file=sys.stderr)
    st, cs = read_sub(args.sub)
    print(f"  {len(st)} submarkets", file=sys.stderr)

    print("Computing derived growth fields...", file=sys.stderr)
    add_growth(mt, periods)

    # last actual period = last period label with no QTD/EST suffix that is
    # immediately followed (in period order) by a QTD/EST label, i.e. the
    # newest plain quarter before the QTD boundary.
    qtd = [p for p in periods if "QTD" in p]
    if qtd:
        boundary_q = qtd[0].replace(" QTD", "")   # e.g. "2026 Q3"
        bi = periods.index(qtd[0])
        # last actual is the plain period just before the QTD entry
        lp = periods[bi - 1]
    else:
        lp = [p for p in periods if "EST" not in p and "QTD" not in p][-1]
    print(f"  last actual (lp) = {lp}", file=sys.stderr)

    print("Building national aggregate...", file=sys.stderr)
    nt = build_national(mt, periods)

    print("Building metro list snapshot...", file=sys.stderr)
    ml = build_ml(mt, cm, lp)

    D = OrderedDict()
    D["p"] = periods
    D["lp"] = lp
    D["ml"] = ml
    D["mt"] = mt
    D["st"] = st
    D["cm"] = cm
    D["cs"] = cs
    D["nt"] = nt

    blob = json.dumps(D, separators=(",", ":"), ensure_ascii=False)
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write(f"var {args.varname} = {blob};\n")
    print(f"Wrote {args.out} ({len(blob)//1024//1024} MB)", file=sys.stderr)

    meta = {"lp": lp, "asof": None, "periods": len(periods),
            "metros": len(mt), "submarkets": len(st),
            "first": periods[0], "last": periods[-1]}
    with open(args.out.rsplit(".", 1)[0] + ".meta.json", "w") as fh:
        json.dump(meta, fh, indent=2)
    print("Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
